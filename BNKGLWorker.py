#! python3
# BNKGLWorker.py - A program that will compare values from the GL and
# bank statement for matches.

import os
import shutil
import configparser
import re
import PyPDF2
import openpyxl
from string import ascii_letters

def MainScript():

	config_dir = os.path.expanduser("~") + "/BNKGLWorker/config"
	config = config_dir + "/user_config.ini"
	os.chdir(config_dir)

	#------Not sure if I need this or not-------
	#if not os.path.isfile(config):
	#    os.makedirs(config_dir, exist_ok=True)
	#    ConfigSetup()

	user_config = configparser.ConfigParser()
	user_config.read(config)


	def bankStmntPDFtoExcelConversion():
		cashAmountsRegex = re.compile(r"\d+?,?\d+\.\d{2}\-?")
		cashAmounts = []
		pagesinPDF = bankStatementPDFReader.numPages
		for i in range(pagesinPDF):
			pageObj = bankStatementPDFReader.getPage(i)
			textFromPage = pageObj.extractText()
			textFromPage.replace(',','')
			for i in cashAmountsRegex.findall(textFromPage):
				cashAmounts.append(i)
		for i in range(len(cashAmounts)):
			if '-' in cashAmounts[i]:
				cashAmounts[i] = '-' + cashAmounts[i][:-1]
			sheetBnkStmnt['C' + str(3 + i)] = cashAmounts[i]


	def extractGLAmountsfromExcel():
		for rowOfGLAmounts in sheetGL[sheetGLStartingCell : sheetGLStartingCell[0] + str(sheetGL.max_row)]:
			for amount in rowOfGLAmounts:
				amountsOnGL.append(amount.value)



	def extractBankStmntAmountsfromExcel():
		for rowOfBnkStmntAmounts in sheetBnkStmnt['C3' : 'C' + str(sheetBnkStmnt.max_row)]:
			for amount in rowOfBnkStmntAmounts:
				amount = amount.value.replace(",", "")
				amountsOnBnkStmnt.append(amount)


	def exactMatchTester():
		for i in range(len(amountsOnGL)):
			for j in range(len(amountsOnBnkStmnt)):
				try:
					if float(amountsOnGL[i]) == float(amountsOnBnkStmnt[j]):
						sheetBnkStmnt['D' + str(3 + j)] = 'Perfect Match!'
						sheetGL[rLetter + str(int(sheetGLStartingCell[1:]) + i)] = 'Perfect Match!'
				except:
					continue


	def alternativeMatchTester():
		for i in range(len(amountsOnGL)):
			for j in range(len(amountsOnBnkStmnt)):
				for k in range(len(amountsOnGL)):
					try:
						if float(amountsOnGL[i]) + float(amountsOnGL[k]) == float(amountsOnBnkStmnt[j]):
							sheetBnkStmnt['D' + str(3 + j)] = 'Possible match with ' + str(sheetGLStartingCell[0] + str(i) + ' and ' + str(k))
							sheetGL[rLetter + str(int(sheetGLStartingCell[1:]) + i)] = 'Possible match with ' + str(sheetGLStartingCell[0]) + str(k)	
					except:
						continue 


	def saveandCloseExcelBooks():
		bankStatementExcel.save(user_config['file_inputs']['bankStatementExcel'])
		bankRecExcel.save(user_config['file_inputs']['bankRecExcel'])


	os.chdir(os.path.expanduser("~") + "/BNKGLWorker")
	bankStatementPDF = open(user_config['file_inputs']['bankStatementPDF'], 'rb')
	bankStatementPDFReader = PyPDF2.PdfFileReader(bankStatementPDF)

	bankRecExcel = openpyxl.load_workbook(user_config['file_inputs']['bankRecExcel'])
	bankStatementExcel = openpyxl.Workbook()

	sheetGL = bankRecExcel[user_config['excel_settings']['sheetGL']]
	sheetGLStartingCell = user_config['excel_settings']['startingCell']
	sheetBnkStmnt = bankStatementExcel['Sheet']

	amountsOnGL = []
	amountsOnBnkStmnt = []
	rLetter = ascii_letters[ascii_letters.index(sheetGLStartingCell[0]) + 1]
		
	bankStmntPDFtoExcelConversion()
	extractGLAmountsfromExcel()
	extractBankStmntAmountsfromExcel()
	alternativeMatchTester()
	exactMatchTester()
	saveandCloseExcelBooks()
