#! python3
# this is a program that I use to reconcile account balances as an accountant.

import openpyxl
import os

matchingAmounts = []
matchingDates = []
matchingDescriptions = []
# These are for storing data for the outputToExcel() function. 

def main():
	wb = openpyxl.load_workbook(os.path.expanduser("~") + "/Book.xlsx") 
	sheet = wb.active
	rowCount = sheet.max_row
	amounts = []
	totalAmount = 0
	matches = 0
	
	for rowOfCells in sheet['H2':'H' + str(rowCount)]:
	    for i in rowOfCells:
	        for j in range(4):
	            sheet['H' + str(rowCount - j)] = 0
	        if i.value == None:
	            amounts.append(0)
	        else:
	            totalAmount = totalAmount + i.value
	            amounts.append(i.value)
	
	for i in range(len(amounts)):
	    doesAmountZero = False
	    for j in range(len(amounts)):
	        if amounts[i] + amounts[j] == 0:
	            amounts[i] = 0
	            amounts[j] = 0
	            doesAmountZero = True
	    if doesAmountZero == False:
	        matches = matches + sheet['H' + str(i + 2)].value
	        matchingAmounts.append(amounts[i])
	        matchingDates.append(sheet['F' + str(i + 2)].value)
	        matchingDescriptions.append(sheet['G' + str(i + 2)].value)
	        print(str(amounts[i]) +
	              ' - ' +
	              str(sheet['F' + str(i + 2)].value)[:7] +
	              '  -  -  -  ' +
	              str(sheet['G' + str(i + 2)].value))
	        
	print('The sum of amounts outstanding is ' + str(matches))
	print('Sum of column H (all values): ' + str(totalAmount))
	if int(matches) != int(totalAmount):
	    print('Matches and total are different; check excel sheet.')
	if int(matches) == int(totalAmount):
	    print('Matches and total are equal!')
	print('Input a name to save output as an Excel sheet (input nothing to continue without saving.)')
	sheetName = input()
	return sheetName

def outputToExcel():
	wb = openpyxl.load_workbook("REDACTED")
	wb.create_sheet(sheetName)
	sheet = wb[sheetName]
	for i in range(0, len(matchingAmounts)):
		sheet["A" + str(i+1)].value = matchingAmounts[i]
		sheet["B" + str(i+1)].value = matchingDates[i]
		sheet["C" + str(i+1)].value = matchingDescriptions[i]

	wb.save("REDACTED")




while __name__ == '__main__':
	print('press any key to begin the program.')
	input()
	sheetName = main()
	if sheetName == '':
		continue
	else:
		outputToExcel()
